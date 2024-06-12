from decimal import Decimal
from distutils.log import debug
from itertools import product
from multiprocessing import context
from django.http import HttpResponseNotFound, HttpResponseRedirect
from django.shortcuts import render, redirect
from django.views.decorators.http import require_POST

from admin.forms import AutoFieldOptionsForm, ComboForm, ConstructorCategoryForm, CustomCodeForm, DeliveryForm, DeliveryTimePriceForm, DopItemsForm, FontForm, FoodConstructorForm, ImageForm, IngridientsForm, IntegrationsForm, LoyaltyCardForm, LoyaltyCardSettingsForm, LoyaltyCardStatusForm, OrderStatusForm, PageItemForm, ProductSaleForm, RelatedProductsForm, CouponForm, CategoryForm, CharGroupForm, CharNameForm, ColorsForm, OptionTypeForm, AlfaBankForm, PayKeeperForm, PaymentForm, PickupAreasForm, PostBlockForm, ProductCharForm, ProductForm, ManufacturerForm, ProductImageForm, ProductOptionForm, RecaptchaSettingsForm, ReviewsForm, SetupForm, EmailSettingsForm, ShopSetupForm, SubdomainsForm, ThemeSettingsForm, BlogCategoryForm, PostForm, SliderSetupForm, SliderForm, PageForm, OrderForm, BlogSetupForm, TinkoffForm, UserForm, UserRightsForm, WorksdayForm, YookassaForm, PayMethodForm
from coupons.models import Coupon
from home.models import Page, PageItem, PlaceImages, Slider, SliderSetup, Reviews
from accounts.models import LoyaltyCard, LoyaltyCardSettings, LoyaltyCardStatus, UserProfile, UserRigts
from integrations.models import Integrations
from subdomains.models import Subdomain
from delivery.models import Delivery

from orders.models import Order, OrderStatus, OrderView
from shop.models import AutoFieldOptions, Category, CharGroup, CharName, DeliveryTimePrice, DopItems, Manufacturer, OptionImage, PayMethod, PickupAreas, Product, OptionType, ProductChar, ProductImage, ProductOption, ProductSale, ShopSetup, WorkDay, FoodConstructor, ConstructorCategory, Ingridients
from setup.models import BaseSettings, Colors, CustomCode, EmailSettings, Fonts, RecaptchaSettings, ThemeSettings
from pay.models import PayKeeper, PaymentSet, Tinkoff, Yookassa, AlfaBank
from blog.models import BlogCategory, BlogSetup, Post, PostBlock

import subprocess
from main.settings import RESET_FILE

from django.contrib.auth import get_user_model
User = get_user_model()

from django.db.models import Q

from django.contrib.auth.decorators import user_passes_test

from django.db.models import Sum
import decimal
from sms.views import send_sms

from functools import wraps


def check_user_rights(required_rights):
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            # Проверяем, авторизован ли пользователь
            if not request.user.is_authenticated:
                return redirect(f'/accounts/login/?next={request.path}')  # Перенаправляем на страницу авторизации с сохранением адреса

            # Проверяем, является ли пользователь суперпользователем
            if request.user.is_superuser:
                return view_func(request, *args, **kwargs)

            # Проверяем, есть ли у пользователя необходимые права доступа
            user_rights = request.user.rights
            if not all(getattr(user_rights, right, False) for right in required_rights):
                return render(request, 'access_denied.html')  # Выводим страницу с сообщением о запрете доступа

            return view_func(request, *args, **kwargs)

        return _wrapped_view
    return decorator



@check_user_rights(['add_workers'])
def get_workers(request):
    
    context = {
        'users': User.objects.all()
    }
    
    return render(request, 'workers/workers.html', context)


@check_user_rights(['add_workers'])
def add_user(request):
    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)  # Сохраняем форму, но не в базу данных
            user.set_password(form.cleaned_data['password'])  # Хешируем пароль
            user.save()  # Теперь сохраняем пользователя с зашифрованным паролем
            user_rights = UserRigts.objects.create(user=user)
            return redirect('get_workers')
    else:
        form = UserForm()
    return render(request, 'workers/add_user.html', {'form': form})


@check_user_rights(['add_workers'])
def delete_user(request, user_id):
    
    user = User.objects.get(id=user_id)
    user.delete()
    return redirect('get_workers')


@check_user_rights(['add_workers'])
def edit_user(request, user_id):
    user = User.objects.get(id=user_id)
    if request.method == 'POST':
        form = UserForm(request.POST, instance=user)
        if form.is_valid():
            user = form.save(commit=False)  # Сохраняем форму, но не в базу данных
            user.set_password(form.cleaned_data['password'])  # Хешируем пароль
            user.save()  # Теперь сохраняем пользователя с зашифрованным паролем
            return redirect('get_workers')
    else:
        form = UserForm(instance=user)
    return render(request, 'workers/edit_user.html', {'form': form})



@check_user_rights(['add_workers'])
def edit_user_rights(request, user_id):
    user = User.objects.get(id=user_id)
    user_rights = UserRigts.objects.get(user=user)
    

    if request.method == 'POST':
        form = UserRightsForm(request.POST, instance=user_rights)
        if form.is_valid():
            user = form.save(commit=False)  # Сохраняем форму, но не в базу данных
            user.save()  # Теперь сохраняем пользователя с зашифрованным паролем
            return redirect('get_workers')

    
        else:

            return render(request, 'workers/edit_user_rights.html', {'form': form})

    

    form = UserRightsForm(instance=user_rights)
    



    return render(request, 'workers/edit_user_rights.html', {'form': form})



# Сессия с хранением состояния сайдбара в админке
@check_user_rights(['view_orders'])
def sidebar_show(request): 
   
    request.session['sidebar'] = 'True' 
    
    return redirect('admin')

@check_user_rights(['view_orders'])
def sidebar_hide(request): 
    
    request.session['sidebar'] = 'False' 
    return redirect('admin')


@check_user_rights(['view_orders'])
def faq(request):
    
    context = {

    }
    
    return render(request, 'faq.html', context)





@check_user_rights(['view_all_statictic'])
def admin(request):

    try:
        setup = BaseSettings.objects.get()
    except:
        setup = BaseSettings.objects.create()
        setup.save()

    try:
        shop_setup = ShopSetup.objects.get()
    except:
        shop_setup = ShopSetup.objects.create()
        shop_setup.save()

    try:
        blog_setup = BlogSetup.objects.get()
    except:
        blog_setup = BlogSetup.objects.create()
        blog_setup.save()
    try:
        email = EmailSettings.objects.get()
    except:
        email = EmailSettings.objects.create()
        email.save()
    try:
        theme = ThemeSettings.objects.get()
    except:
        theme = ThemeSettings.objects.create()
        theme.save()
    try:
        colors = Colors.objects.get()
    except:
        colors = Colors.objects.create()
        colors.save()
    try:
        slider_setup = SliderSetup.objects.get()
    except:
        slider_setup = SliderSetup.objects.create()
        slider_setup.save()

    try:
        lo_settings = LoyaltyCardSettings.objects.get()
    except:
        lo_settings = LoyaltyCardSettings.objects.create()
        lo_settings.save()


    products = Product.objects.all().count()
    orders = Order.objects.all().count()
    sales = Order.objects.all().aggregate(Sum('summ'))

    summ = sales['summ__sum']

    users = User.objects.all().exclude(is_staff=True)
    
    clients_reg = UserProfile.objects.filter(user__in=users).count()
    client_no_reg = UserProfile.objects.filter(user=None).count()

    clients = clients_reg+client_no_reg

    products = Product.objects.all().count()
    
    context = {
        'products': products,
        'orders': orders,
        'summ': summ,
        'clients': clients,
        

    }

    return render(request, 'pages/index.html', context)

@user_passes_test(lambda u: u.is_superuser)
def general_settings_block(request):

    if request.method == 'POST':
        setup = BaseSettings.objects.get()
        block=request.POST.get('block')
        if block == 'on':
            setup.block = True
            
        else:
            setup.block = False

        
        setup.save()
        return redirect('general_settings')
    

@check_user_rights(['change_settings'])
def general_settings(request):
    setup = BaseSettings.objects.get()

    # Сохранение основных настроек
    if request.method == 'POST':
        new_form = SetupForm(request.POST, request.FILES, instance=setup)
        if new_form.is_valid():
            new_form.save()

            subprocess.call(["touch", RESET_FILE])

            return redirect ('general_settings')

    # Заполнение форм значениями, для отображения уже сохраненных настроек
    
    email = EmailSettings.objects.get()
    
    
    form = SetupForm(instance=setup)
    email_form = EmailSettingsForm(instance=email)
    
    context = {
        'form': form,
        'setup': setup,
        'email_form': email_form,
        
    }
    return render(request, 'settings/general_settings.html', context)


# Способы оплаты
@check_user_rights(['add_pay_systems'])
def add_pay_method(request):
    if request.method == 'POST':
        form = PayMethodForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('shop_settings')
        else:
            return render(request, 'shop/pay_method/add_pay_method.html', {'form':form})
    form = PayMethodForm()
    context = {
        'form': form
    }
    return render(request, 'shop/pay_method/add_pay_method.html', context)


@check_user_rights(['add_pay_systems'])
def edit_pay_method(request, pk):
    method = PayMethod.objects.get(id=pk)
    if request.method == 'POST':
        form = PayMethodForm(request.POST, request.FILES, instance=method)
        if form.is_valid():
            form.save()
            return redirect('shop_settings')
        else:
            return render(request, 'shop/pay_method/add_pay_method.html', {'form':form})
    form = PayMethodForm(instance=method)
    context = {
        'form': form
    }
    return render(request, 'shop/pay_method/add_pay_method.html', context)

@check_user_rights(['add_pay_systems'])
def delete_pay_method(request, pk):
    method = PayMethod.objects.get(id=pk)
    method.delete()
    return redirect('shop_settings')


# Настроки платежей
@check_user_rights(['add_pay_systems'])
def admin_payments(request):

    # payment = PaymentSet.objects.get()
    # PaymentSet.delete()

    if request.method == 'POST':

        form = PaymentForm(request.POST)
        if form.is_valid():
            form.save()
            subprocess.call(["touch", RESET_FILE])
            return redirect('admin_payments')
        else:
            return render(request, 'settings/admin_payments.html', {'form':form})


    try:
        payment = PaymentSet.objects.get()
        form = PaymentForm(instance=payment)

    except:
        payment = None
        form = PaymentForm()


    try:
        yookassa = Yookassa.objects.get()
        yookassa_form = YookassaForm(instance=yookassa)
    except:
        yookassa_form = YookassaForm()


    try:
        alfabank = AlfaBank.objects.get()
        alfabank_form = AlfaBankForm(instance=alfabank)
    except:
        alfabank_form = AlfaBankForm()


    try:
        paykeeper = PayKeeper.objects.get()
        paykeeper_form = PayKeeperForm(instance=paykeeper)
    except:
        paykeeper_form = PayKeeperForm()

    try:
        tinkoff = Tinkoff.objects.get()
        tinkoff_form = TinkoffForm(instance=tinkoff)
    except:
        tinkoff_form = TinkoffForm()


    context = {
        'payment': payment,
        'form': form,
        'yookassa_form':yookassa_form,
        'alfabank_form': alfabank_form,
        'paykeeper_form': paykeeper_form,
        'tinkoff_form': tinkoff_form

    }

    return render(request, 'settings/admin_payments.html', context)


@check_user_rights(['add_pay_systems'])
def yookassa_save(request):

    if request.method == 'POST':
        yookassa_form = YookassaForm(request.POST)
        if yookassa_form.is_valid():
            yookassa_form.save()
            subprocess.call(["touch", RESET_FILE])
            return redirect('admin_payments')

        else:
            return redirect('admin_payments')


    else:
        return redirect('admin_payments')


@check_user_rights(['add_pay_systems'])
def alfabank_save(request):

    if request.method == 'POST':
        alfabank_form = AlfaBankForm(request.POST)
        if alfabank_form.is_valid():
            alfabank_form.save()
            subprocess.call(["touch", RESET_FILE])
            return redirect('admin_payments')

        else:
            return redirect('admin_payments')


    else:
        return redirect('admin_payments')


@check_user_rights(['add_pay_systems'])
def paykeeper_save(request):

    if request.method == 'POST':
        paykeeper_form = PayKeeperForm(request.POST)
        if paykeeper_form.is_valid():
            paykeeper_form.save()
            subprocess.call(["touch", RESET_FILE])
            return redirect('admin_payments')

        else:
            return redirect('admin_payments')


    else:
        return redirect('admin_payments')
    

@check_user_rights(['add_pay_systems'])
def tinkoff_save(request):

    if request.method == 'POST':
        paykeeper_form = TinkoffForm(request.POST)
        if paykeeper_form.is_valid():
            paykeeper_form.save()
            subprocess.call(["touch", RESET_FILE])
            return redirect('admin_payments')

        else:
            return redirect('admin_payments')


    else:
        return redirect('admin_payments')
    


# Настройки почты POST
@user_passes_test(lambda u: u.is_superuser)
def email_settings(request):
    if request.method == 'POST':
        form = EmailSettingsForm(request.POST)
        if form.is_valid():
            form.save()

            subprocess.call(["touch", RESET_FILE])
            return redirect('general_settings')


# Настройки recaptcha POST
@user_passes_test(lambda u: u.is_superuser)
def recaptcha_settings(request):
    if request.method == 'POST':
        form = RecaptchaSettingsForm(request.POST)
        if form.is_valid():
            form.save()
            subprocess.call(["touch", RESET_FILE])
            return redirect('general_settings')


# Настройки кастомных кодов POST/GET
@check_user_rights(['add_codes'])
def codes_settings(request):
    if request.method == 'POST':
        form = CustomCodeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('codes_settings')
        else:
            return render(request, 'settings/codes_settings.html', { 'form': form })
    codes = CustomCode.objects.all()
    form = CustomCodeForm()
    context = {
        'codes': codes,
        'form': form
    }
    return render(request, 'settings/codes_settings.html', context)


@check_user_rights(['add_codes'])
def codes_settings_edit(request, pk):
    codes = CustomCode.objects.get(pk=pk)
    if request.method == 'POST':
        form = CustomCodeForm(request.POST, instance=codes)
        if form.is_valid():
            form.save()
            return redirect('codes_settings')
        else:
            return render(request, 'settings/codes_edit.html', { 'form': form })
    
    form = CustomCodeForm(instance=codes)
    context = {
        'form': form
    }


    return render(request, 'settings/codes_edit.html', context)



@check_user_rights(['add_codes'])
def codes_settings_delete(request, pk):
    codes = CustomCode.objects.get(pk=pk)
    codes.delete()
    return redirect('codes_settings')


# Настройки цвета
@check_user_rights(['change_design'])
def color_settings(request):
    try:
        color = Colors.objects.get()
        form = ColorsForm(instance=color)
    except:
        form = ColorsForm()

    if request.method == 'POST':
        form = ColorsForm(request.POST, request.FILES, instance=color)
        if form.is_valid():
            form.save()

        else:
            return render(request, 'settings/color_settings.html', {'form': form})

        return redirect('color_settings')
    
    context = {
        'form': form
    }

    return render(request, 'settings/color_settings.html', context)


# Настройка темы
@check_user_rights(['change_design'])
def theme_settings(request):
    if request.method == 'POST':

        form = ThemeSettingsForm(request.POST)
        
        if form.is_valid():
            form.save()
            subprocess.call(["touch", RESET_FILE])
        return redirect('theme_settings')

    try:
        theme = ThemeSettings.objects.get()
        form = ThemeSettingsForm(instance=theme)
    except:
        form = ThemeSettingsForm()
    
    try:
        font = Fonts.objects.get()
        font_form = FontForm(instance=font)
    except:
        font_form = FontForm()
    
    context = {
        'form': form,
        'font_form': font_form
    }

    return render(request, 'settings/theme_settings.html', context)

@check_user_rights(['change_design'])
def font_settings(request):
    
    if request.method == 'POST':

        form = FontForm(request.POST)
        if form.is_valid():
            form.save()
            subprocess.call(["touch", RESET_FILE])
        else:
            print(form.errors)

        return redirect('theme_settings')





# !!! МАРКЕТИНГ !!!
# Промокоды
@check_user_rights(['add_propmo'])
def admin_promo(request):

    context = {
        'coupons': Coupon.objects.all().order_by('valid_to')
    }

    return render(request, 'marketing/coupons/promo.html', context)

@check_user_rights(['add_propmo'])
def promo_add(request):

    if request.method == 'POST':
        form = CouponForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('admin_promo')

    form = CouponForm()
    context = {
        'form': form
    }

    return render(request, 'marketing/coupons/promo_add.html', context)

@check_user_rights(['add_propmo'])
def promo_edit(request, pk):
    coupon = Coupon.objects.get(id=pk)

    if request.method == 'POST':
        form = CouponForm(request.POST, instance=coupon)
        if form.is_valid():
            form.save()

            return redirect('admin_promo')

        else:
            return render(request, 'marketing/coupons/promo_edit.html', {'form':form})

    form = CouponForm(instance=coupon)
    context = {
        'form': form,
        'coupon': coupon
    }

    return render(request, 'marketing/coupons/promo_edit.html', context)

@check_user_rights(['add_propmo'])
def promo_delete(request, pk):
    coupon = Coupon.objects.get(id=pk)
    coupon.delete()
    return redirect('admin_promo')




# Карты лояльности
@check_user_rights(['redact_loyal_cart'])
def card_settings(request):

    settings = LoyaltyCardSettings.objects.get()

    if request.method == 'POST':
        if settings:
            form = LoyaltyCardSettingsForm(request.POST, instance=settings)
        else:
            form = LoyaltyCardSettingsForm(request.POST)

        if form.is_valid():
            form.save()
            
            return redirect('admin_card')

        else:
            
            return render(request, 'marketing/loyalty_card/card_settings.html', {'form':form})

    if settings:
        form = LoyaltyCardSettingsForm(instance=settings)
    else:
        form = LoyaltyCardSettingsForm()

    context = {

        'form': form

    }
    return render(request, 'marketing/loyalty_card/card_settings.html', context)


@check_user_rights(['redact_loyal_cart'])
def admin_card(request):


    if request.method == 'POST':
        form = LoyaltyCardStatusForm(request.POST)

        if form.is_valid():
            form.save()

            return redirect('admin_card')
        
        else:
            
            return render(request, 'marketing/loyalty_card/admin_card.html', {'form': form})




    form = LoyaltyCardStatusForm()

    card_statuses = LoyaltyCardStatus.objects.all()
    user_cards = LoyaltyCard.objects.all()

    try:
        setting = LoyaltyCardSettings.objects.get()
    except:
        setting = LoyaltyCardSettings.objects.create()

    context = {

        'card_statuses': card_statuses,
        'user_cards': user_cards,
        'form': form,
        'setting': setting,


    }
    return render(request, 'marketing/loyalty_card/admin_card.html', context)


@check_user_rights(['redact_loyal_cart'])
def loyalty_card_status_edit(request, pk):
    card_status = LoyaltyCardStatus.objects.get(id=pk)

    if request.method == 'POST':
        form = LoyaltyCardStatusForm(request.POST, instance=card_status)
        if form.is_valid():
            form.save()

            return redirect('admin_card')
        
        else:
            return render(request, 'marketing/loyalty_card/loyalty_card_status_edit.html', {'form': form})




    form = LoyaltyCardStatusForm(instance=card_status)

    context = {
        'form': form
    }

    return render(request, 'marketing/loyalty_card/loyalty_card_status_edit.html', context)


    


@check_user_rights(['redact_loyal_cart'])
def loyalty_card_status_delete(request, pk):
    status = LoyaltyCardStatus.objects.get(id=pk)
    status.delete()

    return redirect('admin_card')





@check_user_rights(['redact_loyal_cart'])
def card_add(request):

    context = {

    }


    return render(request, 'marketing/loyalty_card/card_add.html', context)

@check_user_rights(['redact_loyal_cart'])
def card_edit(request, pk):

    card = LoyaltyCard.objects.get(id=pk)

    if request.method == 'POST':
        form = LoyaltyCardForm(request.POST, instance=card)
        if form.is_valid():
            form.save()
            return redirect('admin_card')
        else:

            return render(request, 'marketing/loyalty_card/card_edit.html', {'form':form})
    
    
    form = LoyaltyCardForm(instance=card)
    
    context = {
        'form': form
    }

    return render(request, 'marketing/loyalty_card/card_edit.html', context)

@check_user_rights(['redact_loyal_cart'])
def card_delete(request, pk):
    
    
    card = LoyaltyCard.objects.get(id=pk)
    card.delete()

    return redirect('admin_card')




# !!! end МАРКЕТИНГ !!!




# !!! Продажи !!!

# Заказы
@check_user_rights(['view_orders'])
def admin_order(request):
    orders = Order.objects.all().order_by('-created')
    pay = PaymentSet.objects.filter(status=True, name='alfabank').first()
    
    context = {
        'orders': orders,
        'pay': pay
        
    }


    return render(request, 'order/admin_order.html', context)

@check_user_rights(['view_orders'])
def order_detail(request, pk):
    order = Order.objects.get(id=pk)
    form = OrderForm(instance=order)

    user = request.user

    try:
        order_view = OrderView.objects.get(order=order, user=user)
    except:
        order_view = OrderView.objects.create(order=order, user=user)

    context = {
        'order': order,
        'form': form,
    }


    return render(request, 'order/order_detail.html', context)


@check_user_rights(['view_orders'])
def order_view_all(request):
    orders = Order.objects.all().order_by('-created')
    for order in orders:
        try:
            order_view = OrderView.objects.get(order=order, user=request.user)
        except:
            order_view = OrderView.objects.create(order=order, user=request.user)


    return redirect('admin_order')

from delivery.yandex_eda import yandex_create_order
from orders.telegram import send_message

@check_user_rights(['view_orders'])
def order_status_change(request, pk):
    order = Order.objects.get(id=pk)
    order_prev_status = order.status
    loyalty_settings = LoyaltyCardSettings.objects.get()

    if request.method == 'POST':
        form = OrderForm(request.POST, instance=order)
        
        try:
            
            if form.is_valid():
                status = form.cleaned_data['status']


                if status == 'Готов к доставке':

                    
                    yandex_create_order(order)



                if loyalty_settings.active == True:
                    

                    user = order.user_pr
                    card = LoyaltyCard.objects.get(user=user)
                    user_orders = Order.objects.filter(user_pr=user)

                    card = LoyaltyCard.objects.get(user=user)

                    
                    order_count = user_orders.count()

                    enable_add_balls_after_first_order = loyalty_settings.enable_add_balls_after_first_order
                    balls_after_first_order = loyalty_settings.balls_after_first_order
                    first_order_summ_for_add_balls = loyalty_settings.first_order_summ_for_add_balls
                    send_sms_status = loyalty_settings.send_sms
                    sms_text = loyalty_settings.sms_text
                    balls_summ = 0



                    if status == 'Выполнен':
                        
                        if order_count == 1 and enable_add_balls_after_first_order and order.summ >= first_order_summ_for_add_balls:
                            card.balls = card.balls + balls_after_first_order
                            card.summ = Decimal(card.summ) + (Decimal(order.summ) - Decimal(order.delivery_price))
                            balls_summ = balls_after_first_order


                    
                        else:
                            if loyalty_settings.status_up == True:
                                card.summ = Decimal(card.summ) + (Decimal(order.summ) - Decimal(order.delivery_price))
                                card.balls = card.balls + (((Decimal(order.summ) - Decimal(order.delivery_price)) / 100) * card.status().percent_up).quantize(Decimal("1"), decimal.ROUND_DOWN) 
                                balls_summ = (((Decimal(order.summ) - Decimal(order.delivery_price)) / 100) * card.status().percent_up).quantize(Decimal("1"), decimal.ROUND_DOWN) 



                        if send_sms_status == True:
                            try:
                                site_name = BaseSettings.objects.get().name
                            except:
                                site_name = ''    
                            
                            if site_name:
                                sms_text = sms_text.replace('{balls}', str(balls_summ)).replace('{sitename}', site_name)
                            else:
                                sms_text = sms_text.replace('{balls}', str(balls_summ)).replace('- {sitename}', '')
                            phone = order.phone
                            if balls_summ != 0:
                                send_sms(sms_text, phone)
                            

                    if status == 'Отказ':
                        
                        if order_prev_status == 'Выполнен':
                        
                            if order_count == 1 and enable_add_balls_after_first_order and order.summ >= first_order_summ_for_add_balls:
                                card.balls = card.balls - balls_after_first_order
                                card.summ = Decimal(card.summ) - (Decimal(order.summ) - Decimal(order.delivery_price))
                            else:
                                if loyalty_settings.status_up == True:
                                    card.summ = Decimal(card.summ) - (Decimal(order.summ) - Decimal(order.delivery_price))
                                    card.balls = card.balls - (((Decimal(order.summ) - Decimal(order.delivery_price)) / 100) * card.status().percent_up).quantize(Decimal("1"), decimal.ROUND_DOWN) 


                    card.save()

                form.save()

                return redirect('order_detail', order.id)
            
        
        except Exception as e:
            telegram_bot = '5953442472:AAHsgzGdcVrnuJnb0FnDWJ4nrPdDT59YNOE'
            telegram_group = '-1002079435900'

            send_message(telegram_bot, telegram_group, f'ОШИБКА при изменении статуса: {e}')
            

            print(e)


@check_user_rights(['view_orders'])
def order_delete(request, pk):
    order = Order.objects.get(id=pk)
    order.delete()
    return redirect('admin_order')

@check_user_rights(['view_orders'])
def order_status(request):
    order_statuses = OrderStatus.objects.all().order_by('sort')
    context = {
        'order_statuses': order_statuses
    }
    return render(request, 'order/order_status.html', context)


@check_user_rights(['view_orders'])
def add_order_status(request):
    if request.method == 'POST':
        form = OrderStatusForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            
            return redirect('admin_order')
    else:
        form = OrderStatusForm()
    context = {
        'form': form
    }
    return render(request, 'order/add_order_status.html', context)



@check_user_rights(['view_orders'])
def edit_order_status(request, pk):
    order_status = OrderStatus.objects.get(id=pk)
    if request.method == 'POST':
        form = OrderStatusForm(request.POST, request.FILES, instance=order_status)
        if form.is_valid():
            form.save()
            
            return redirect('admin_order')
    else:
        form = OrderStatusForm(instance=order_status)
    context = {
        'form': form
    }
    return render(request, 'order/add_order_status.html', context)


@check_user_rights(['view_orders'])
def delete_order_status(request, pk):
    order_status = OrderStatus.objects.get(id=pk)
    order_status.delete()
    
    return redirect('admin_order')


# !!! end Продажи !!!




# !!! Загрузка файлла с зонами доставки !!!
import json
import os

@check_user_rights(['change_shop_settings'])
def zone_file(request):

    if request.method == 'POST':
        json_file = request.FILES['file']
        deliverys = json.load(json_file)
        new_file = {
            "deliverys": [
            ]
        }
        for d in deliverys['features']:
            if d['properties']['description'] == '0':
                hintContent = 'Зона бесплатной доставки'
                balloonContent = 'Бесплатная доставка'
                balloonContentHeader = 'Зона бесплатной доставки'
                balloonContentBody = 'Стоимость доставки'
                balloonContentFooter = d['properties']['description'] + ' рублей'
            else:
                hintContent = 'Зона платной доставки'
                balloonContent = 'Платная доставка'
                balloonContentHeader = 'Зона платной доставки'
                try:
                    balloonContentBody = 'Стоимость доставки - ' + d['properties']['description'].split(':')[0] + ' рублей'

                    try:
                        min_delivery = d['properties']['description'].split(':')[2]
                    except Exception as e:
                        min_delivery = None

                    try:
                        free_delivery = d['properties']['description'].split(':')[1]
                        if free_delivery == '999999' or free_delivery == '0' or free_delivery == 999999:
                            free_delivery = None
                    except Exception as e:
                        free_delivery = None
                        

                    
                    if free_delivery and min_delivery:
                        balloonContentFooter = f'''Бесплатная доставка от - ''' + d['properties']['description'].split(':')[1]  + ''' рублей,
                        минимальная сумма для заказа - ''' + d['properties']['description'].split(':')[2]  + ''' рублей'''

                    elif min_delivery and not free_delivery:
                        balloonContentFooter = f'''Бесплатная доставка - ''' + 'НЕТ'  + ''',
                        минимальная сумма для заказа - ''' + d['properties']['description'].split(':')[2]  + ''' рублей'''

                    elif free_delivery and not min_delivery:
                        balloonContentFooter = f'''Бесплатная доставка от - ''' + d['properties']['description'].split(':')[1]  + ''' рублей'''

                except Exception as e:
                   
                    balloonContentBody = 'Стоимость доставки - ' + d['properties']['description'] + ' рублей'
                    
            coords = []
            for i in d['geometry']['coordinates']:
                for l in i:
                    coords.append(
                        [str(l[1]), str(l[0])]
                    )
            new_file['deliverys'].append({
                'hintContent': hintContent,
                'balloonContent': balloonContent,
                'balloonContentHeader': balloonContentHeader,
                'balloonContentBody': balloonContentBody,
                'balloonContentFooter': balloonContentFooter,
                'coords': coords,
                "fillColor": d['properties']['fill'],
                "strokeColor": d['properties']['stroke'],
                "opacity": d['properties']['fill-opacity'],
            })

        subdomains = Subdomain.objects.all()

        if subdomains.exists():
            from django.core.files.base import ContentFile
            subdomain_id = request.POST['subdomain']
            subdomain = Subdomain.objects.get(id=subdomain_id)
            subdomain.zone_file.save(f'{subdomain.subdomain}.json', ContentFile(json.dumps(new_file)))
            subdomain.save()

        else:

            try:
                with open('../core/libs/delivery.json', 'w', encoding='utf-8') as f:
                    json.dump(new_file, f, ensure_ascii=False, indent=4)
            except:
                with open('core/libs/delivery.json', 'w', encoding='utf-8') as f:
                    json.dump(new_file, f, ensure_ascii=False, indent=4)

        subprocess.call(["touch", RESET_FILE])

        
        return redirect('shop_settings')
# !!! Загрузка файлла с зонами доставки !!!
    
# !!! district_setup !!!

from .get_dictricts import get_file
@check_user_rights(['change_shop_settings'])
def district_setup(request):

    if request.method == 'POST':
        data_rions = request.POST.get('rions')

        get_file(data_rions)

        return redirect('shop_settings')
    
    file_path = "../core/result.geojson"

    if os.path.exists(file_path):
        file = True
    else:
        file = False

    
    
    context = {
        'file': file
    }

    
    subprocess.call(["touch", RESET_FILE])
    return render(request, 'settings/district_setup.html', context)

# !!! district_setup !!!



# !!! Добавить зону доставки !!!
@check_user_rights(['change_shop_settings'])
def add_zone(request):
    if request.method == 'POST':
        form = PickupAreasForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('shop_settings')
        else:
            return render(request, 'shop/zones/add_zone.html', {'form':form})
    form = PickupAreasForm()
    context = {
        'form': form
    }
    return render(request, 'shop/zones/add_zone.html', context)


@check_user_rights(['change_shop_settings'])
def edit_zone(request, pk):
    zone = PickupAreas.objects.get(id=pk)
    if request.method == 'POST':
        form = PickupAreasForm(request.POST, request.FILES, instance=zone)
        if form.is_valid():
            form.save()
            return redirect('shop_settings')
        else:
            return render(request, 'shop/zones/add_zone.html', {'form':form})
    form = PickupAreasForm(instance=zone)
    context = {
        'form': form
    }
    return render(request, 'shop/zones/add_zone.html', context)


@check_user_rights(['change_shop_settings'])
def delete_zone(request, pk):
    zone = PickupAreas.objects.get(id=pk)
    zone.delete()
    return redirect('shop_settings')



# !!! Добавить зону доставки !!!


# !!!!! МАГАЗИН !!!!! 

@check_user_rights(['change_shop_settings'])
def shop_settings(request):
    try:
        shop_setup = ShopSetup.objects.get()
        form = ShopSetupForm(instance=shop_setup)
    except:
        form = ShopSetupForm()

    if request.method == 'POST':
        shop_setup = ShopSetup.objects.get()
        form_new = ShopSetupForm(request.POST, request.FILES, instance=shop_setup)
        if form_new.is_valid():
            form_new.save()
            subprocess.call(["touch", RESET_FILE])

            return redirect('shop_settings')

        else:
            return render(request, 'shop/settings.html', {'form': form})


    items = DopItems.objects.all()


    context = {
        'form': form,
        'zones': PickupAreas.objects.all(),
        'methods': PayMethod.objects.all(),
        'worksdays': WorkDay.objects.all().order_by('day'),
        'items': items

    }

    return render(request, 'shop/settings.html', context)


@check_user_rights(['change_shop_settings'])
def dop_items_add(request):

    if request.method == 'POST':
        form = DopItemsForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('shop_settings')
        else:
            return render(request, 'shop/dop_items/dop_items_add.html', {'form':form})

    form = DopItemsForm()
    

    context = {
        'form': form,
        
    }


    return render(request, 'shop/dop_items/dop_items_add.html', context)


@check_user_rights(['change_shop_settings'])
def dop_items_edit(request, pk):

    if request.method == 'POST':
        item = DopItems.objects.get(id=pk)
        form = DopItemsForm(request.POST, request.FILES, instance=item)
        if form.is_valid():
            form.save()
            return redirect('shop_settings')
        else:
            return render(request, 'shop/dop_items/dop_items_edit.html', {'form':form})

    item = DopItems.objects.get(id=pk)
    form = DopItemsForm(instance=item)
    context = {
        'form': form,
        
    }

    return render(request, 'shop/dop_items/dop_items_edit.html', context)


@check_user_rights(['change_shop_settings'])
def dop_items_delete(request, pk):

    item = DopItems.objects.get(id=pk)
    item.delete()


    return redirect('shop_settings')
    

# Настройка скидок на товары
@check_user_rights(['change_time_sales'])
def admin_sale(request):
    

    sales = ProductSale.objects.all()


    context = {
        'sales': sales
    }

    return render(request, 'shop/sale/sale.html', context)



@check_user_rights(['change_time_sales'])
def add_sale(request):
    if request.method == 'POST':
        form = ProductSaleForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('admin_sale')
        else:
            return render(request, 'shop/sale/add_sale.html', {'form':form})
    form = ProductSaleForm()
    context = {
        'form': form
    }
    return render(request, 'shop/sale/add_sale.html', context)


@check_user_rights(['change_time_sales'])
def edit_sale(request, pk):
    sale = ProductSale.objects.get(id=pk)
    if request.method == 'POST':
        form = ProductSaleForm(request.POST, request.FILES, instance=sale)
        if form.is_valid():
            form.save()
            return redirect('admin_sale')
        else:
            return render(request, 'shop/sale/add_sale.html', {'form':form})
    form = ProductSaleForm(instance=sale)
    context = {
        'form': form
    }
    return render(request, 'shop/sale/add_sale.html', context)

@check_user_rights(['change_time_sales'])
def delete_sale(request, pk):
    sale = ProductSale.objects.get(id=pk)
    sale.delete()
    return redirect('admin_sale')


# Список категорий
@check_user_rights(['add_categorys'])
def admin_category(request):
    q = request.GET.get('q')
    sort = request.GET.getlist('sort')
    try:
        sort_t = sort[0]
    except:
        sort_t = sort
    
    try:
        categorys = Category.objects.filter(Q(name__icontains=q)).order_by(*sort)
    except:
        categorys = Category.objects.all().order_by('sort_order')
    context = {
        # Разрешить поиск на странице
        'search': 'search',
        'categorys': categorys,
        'q': q,
        'sort': sort_t,
    }
    return render(request, 'shop/category/category.html', context)


# Добавление категорий
@check_user_rights(['add_categorys'])
def category_add(request):
    if request.method == 'POST':
        form_new = CategoryForm(request.POST, request.FILES)
        if form_new.is_valid():
            form_new.save()
            return redirect('admin_category')
        else:
            return render(request, 'shop/category/category_add.html', {'form': form_new})
    context = {
        'form': CategoryForm(),
        'categorys': Category.objects.filter()
    }
    return render(request, 'shop/category/category_add.html', context)


# Удаление категорий
@check_user_rights(['add_categorys'])
def category_delete(request, pk):
    category = Category.objects.get(id=pk)
    category.delete()
    return redirect('admin_category')



@check_user_rights(['add_categorys'])
def cat_orderby_edit(request, pk):


    category = Category.objects.get(id=pk)

    if request.method == 'POST':
        order_by = request.POST['order']
        category.sort_order = order_by
        category.save()
    
    
        return redirect('admin_category')


# Редкатирование категорий
@check_user_rights(['add_categorys'])
def category_edit(request, pk):
    cat = Category.objects.get(id=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES, instance=cat)
        if form.is_valid():
            form.save()
            return redirect('admin_category')
        else:
            return render(request, 'shop/category/category_edit.html', {'form': form})


    
    context = {
        'form': CategoryForm(instance=cat),
        'category': cat,
        'categorys': Category.objects.filter().exclude(id=pk)
    }

    return render(request, 'shop/category/category_edit.html', context)



# Производители
@check_user_rights(['add_categorys'])
def admin_manufacturer(request):
    q = request.GET.get('q')
    sort = request.GET.getlist('sort')
    try:
        sort_t = sort[0]
    except:
        sort_t = sort
    try:
        manufacturer = Manufacturer.objects.filter(Q(name__icontains=q)).order_by(*sort)
    except:
        manufacturer = Manufacturer.objects.all().order_by(*sort)
    context = {
        # Разрешить поиск на странице
        'search': 'search',
        'manufacturer': manufacturer,
        'q': q,
        'sort': sort_t,
    }
    return render(request, 'shop/manufacturer/manufacturer.html', context)

# Добавить производителя
@check_user_rights(['add_categorys'])
def manufacturer_add(request):
    if request.method == 'POST':
        form_new = ManufacturerForm(request.POST, request.FILES)
        if form_new.is_valid():
            form_new.save()
            return redirect('admin_manufacturer')
        else:
            return render(request, 'shop/manufacturer/manufacturer_add.html', {'form': form_new})
    form = ManufacturerForm()
    context = {
       'form': form,
    }
    return render(request, 'shop/manufacturer/manufacturer_add.html', context)


# Редактировать производителя
# !!! Некорректное поведение при редкатирвовании картинки, поправить !!!
@check_user_rights(['add_categorys'])
def manufacturer_edit(request, pk):
    manufacturer = Manufacturer.objects.get(id=pk)
    if request.method == 'POST':
        form_new = ManufacturerForm(request.POST, request.FILES, instance=manufacturer)
        if form_new.is_valid():
            form_new.save()
            return redirect('admin_manufacturer')
        else:
            return render(request, 'shop/manufacturer/manufacturer_edit.html', {'form': form_new})
    
    form = ManufacturerForm(instance=manufacturer)
    context = {
       'form': form,
    }
    return render(request, 'shop/manufacturer/manufacturer_edit.html', context)


# Удалить производителя
@check_user_rights(['add_categorys'])
def manufacturer_delete(request, pk):
    manufacturer = Manufacturer.objects.get(id=pk)
    manufacturer.delete()
    return redirect('admin_manufacturer')



# Опции
@check_user_rights(['add_options'])
def admin_option_type(request):
    q = request.GET.get('q')
    sort = request.GET.getlist('sort')
    try:
        sort_t = sort[0]
    except:
        sort_t = sort
    try:
        options = OptionType.objects.filter(Q(name__icontains=q)).order_by(*sort)
    except:
        options = OptionType.objects.all().order_by(*sort)
    context = {
        # Разрешить поиск на странице
        'search': 'search',
        'options': options,
        'q': q,
        'sort': sort_t,
    }
    return render(request, 'shop/option_type/option_type.html', context)


# Добавить опцию
@check_user_rights(['add_options'])
def option_type_add(request):
    if request.method == 'POST':
        form_new = OptionTypeForm(request.POST)
        if form_new.is_valid():
            form_new.save()
            return redirect('admin_option_type')
        else:
            return render(request, 'shop/option_type/option_type_add.html', {'form': form_new})
    form = OptionTypeForm()
    context = {
       'form': form,
    }
    return render(request, 'shop/option_type/option_type_add.html', context)

# Редкатировать опцию
@check_user_rights(['add_options'])
def option_type_edit(request, pk):
    option_type = OptionType.objects.get(id=pk)
    if request.method == 'POST':
        form_new = OptionTypeForm(request.POST, instance=option_type)
        if form_new.is_valid():
            form_new.save()
            return redirect('admin_option_type')
        else:
            return render(request, 'shop/option_type/option_type_edit.html', {'form': form_new})
    
    form = OptionTypeForm(instance=option_type)
    context = {
       'form': form,
    }
    return render(request, 'shop/option_type/option_type_edit.html', context)

# Удалить опцию
@check_user_rights(['add_options'])
def option_type_delete(request, pk):
    option_type = OptionType.objects.get(id=pk)
    option_type.delete()
    return redirect('admin_option_type')




# Автозаполнение для опций

@check_user_rights(['add_options'])
def option_autofield_add(request):
    if request.method == 'POST':
        form_new = AutoFieldOptionsForm(request.POST)
        if form_new.is_valid():
            form_new.save()
            return redirect('admin_option_type')
        else:
            return render(request, 'shop/option_type/option_autofield_add.html', {'form': form_new})

    form = AutoFieldOptionsForm()
    
    context = {
        'form': form
    }

    return render(request, 'shop/option_type/option_autofield_add.html', context)


@check_user_rights(['add_options'])
def option_autofield_edit(request, pk):
    option_autofield = AutoFieldOptions.objects.get(id=pk)
    if request.method == 'POST':
        form_new = AutoFieldOptionsForm(request.POST, instance=option_autofield)
        if form_new.is_valid():
            form_new.save()
            return redirect('admin_option_type')
        else:
            return render(request, 'shop/option_type/option_autofield_edit.html', {'form': form_new})
    
    form = AutoFieldOptionsForm(instance=option_autofield)
    context = {
       'form': form,
    }
    return render(request, 'shop/option_type/option_autofield_edit.html', context)

@check_user_rights(['add_options'])
def option_autofield_delete(request, pk):
    option_autofield = AutoFieldOptions.objects.get(id=pk)
    option_autofield.delete()
    return redirect('admin_option_type')


@check_user_rights(['add_options'])
def option_autofield_detail(request, pk):
    option = OptionType.objects.get(id=pk)
    context = {
        'option': option,
    }
    return render(request, 'shop/option_type/option_autofield_detail.html', context)



# Характеристики
@check_user_rights(['add_chars'])
def admin_char(request):
    context = {
        'groups': CharGroup.objects.all(),
        'chars': CharName.objects.filter(group=None)
    }
    return render(request, 'shop/char/char.html', context)

@check_user_rights(['add_chars'])
def char_group_add(request):
    if request.method == 'POST':
        form_new = CharGroupForm(request.POST)
        if form_new.is_valid():
            form_new.save()
            return redirect('admin_char')
        else:
            return render(request, 'shop/char/char_group_add.html', {'form': form})

    form = CharGroupForm()
    context = {
        'form': form,
    }
    return render(request, 'shop/char/char_group_add.html', context)

@check_user_rights(['add_chars'])
def char_group_edit(request, pk):
    char_group = CharGroup.objects.get(id=pk)
    if request.method == 'POST':
        form_new = CharGroupForm(request.POST, instance=char_group)
        if form_new.is_valid():
            form_new.save()
            return redirect('admin_char')
        else:
            return render(request, 'shop/char/char_group_edit.html', {'form': form})
    form = CharGroupForm(instance=char_group)
    context = {
        'form': form,
    }
    return render(request, 'shop/char/char_group_edit.html', context)

@check_user_rights(['add_chars'])
def char_group_delete(request, pk):
    char_group = CharGroup.objects.get(id=pk)
    char_group.delete()
    return redirect('admin_char')

@check_user_rights(['add_chars'])
def char_add(request):
    if request.method == 'POST':
        form_new = CharNameForm(request.POST)
        if form_new.is_valid():
            form_new.save()
            return redirect('admin_char')
        else:
            return render(request, 'shop/char/char_add.html', {'form': form})

    form = CharNameForm()
    context = {
        'form': form,
    }
    return render(request, 'shop/char/char_add.html', context)

@check_user_rights(['add_chars'])
def char_edit(request, pk):
    char = CharName.objects.get(id=pk)
    if request.method == 'POST':
        form_new = CharNameForm(request.POST, instance=char)
        if form_new.is_valid():
            form_new.save()
            return redirect('admin_char')
        else:
            return render(request, 'shop/char/char_edit.html', {'form': form})

    form = CharNameForm(instance=char)
    context = {
        'form': form,
    }
    return render(request, 'shop/char/char_edit.html', context)

@check_user_rights(['add_chars'])
def char_delete(request, pk):
    char = CharName.objects.get(id=pk)
    char.delete()
    return redirect('admin_char')


# Товары
@check_user_rights(['add_products'])
def admin_product(request):
    q = request.GET.get('q')
    sort = request.GET.getlist('sort')
    try:
        sort_t = sort[0]
    except:
        sort_t = sort
    try:
        product = Product.objects.filter(Q(name__icontains=q)).exclude(related=True).order_by(*sort)
    except:
        product = Product.objects.all().exclude(related=True).order_by(*sort)
    context = {
        # Разрешить поиск на странице
        'search': 'search',
        'products': product,
        'no_cats': Product.objects.filter(parent=None, related=False),
        'q': q,
        'sort': sort_t,
    }


    return render(request, 'shop/product/product.html', context)

# Добавить товар
@check_user_rights(['add_products'])
def product_add(request):
    form = ProductForm()    
    option_form = ProductOptionForm()
    product_char_form = ProductCharForm()
    image_form = ProductImageForm()
    if request.method == 'POST':
        form_new = ProductForm(request.POST, request.FILES)
        if form_new.is_valid():

            

            form_new.save()
            product = Product.objects.get(slug=request.POST['slug'])

            # Изображения

            images = request.FILES.getlist('src')
            
            for image in images:
                img = ProductImage(parent=product, src=image)
                img.save()

            # Опции

            # opt = ProductOption(
            #     parent = product,
            #     option_sku=product.sku,
            #     option_value = product.name,
            #     option_stock = product.stock,
            #     option_price = 0,
            #     option_subtract = product.subtract,
            #     image_status = False,

            # )
            # opt.save()


            options = request.POST.getlist('type')
            option_weight = request.POST.getlist('option_weight')
            option_value = request.POST.getlist('option_value')
            option_sort = request.POST.getlist('option_sort')
            option_price = request.POST.getlist('option_price')

            
            
            image_status = request.POST.getlist('image_status')
            o_count = 0
            for option in options:
                sort = option_sort[o_count]
                if sort == '':
                    sort = 0
                opt = ProductOption(
                    parent = product,
                    type_id = option,
                    option_weight = option_weight[o_count],
                    option_value = option_value[o_count],
                    sort = sort,
                    option_price = Decimal(option_price[o_count]),
                    
                    image_status = image_status[o_count],
                )
                opt.save()
                
                try:
                    images_name = 'option_images-'+str(o_count)
                    option_images = request.FILES.getlist(images_name)
                    for image in option_images:
                        o_image = OptionImage(parent=opt, src=image)
                        o_image.save()
                except:
                    pass

                o_count += 1


            # Характеристики

            char_name = request.POST.getlist('text_name')
            char_value = request.POST.getlist('char_value')
            char_count = 0

            for char in char_name:

                value = char_value[char_count]
                product_char = ProductChar(
                    char_name_id = char,
                    parent = product,
                    char_value = value
                )
                product_char.save()
                char_count += 1


            product.save()

            return redirect('admin_product')
        else:
            return render(request, 'shop/product/product_add.html', {
                'form': form_new,
                'option_form': option_form,
                'product_char_form': product_char_form,
                'image_form': image_form,
                })

    context = {
       'form': form,
       'option_form': option_form,
       'product_char_form': product_char_form,
       'image_form': image_form,

    }
    return render(request, 'shop/product/product_add.html', context)

@check_user_rights(['add_products'])
def product_edit(request, pk):

    product = Product.objects.get(id=pk)
    form = ProductForm(instance=product) 
    option_form = ProductOptionForm()
    product_char_form = ProductCharForm()

    options = ProductOption.objects.filter(parent_id=pk).order_by('sort')
    chars = ProductChar.objects.filter(parent_id=pk)
    images = ProductImage.objects.filter(parent_id=pk)
    all_options = OptionType.objects.all()
    all_chars = CharName.objects.all()
    

    form_new = ProductForm(request.POST, request.FILES, instance=product) 
    if request.method == 'POST':  
        if form_new.is_valid():
            form_new.save()

            # Изображения

            images = request.FILES.getlist('src')
            
            for image in images:
                img = ProductImage(parent=product, src=image)
                img.save()


            # Характеристики 
            char_name = request.POST.getlist('text_name')
            char_value = request.POST.getlist('char_value')
            char_count = 0

            for char in char_name:

                value = char_value[char_count]
                product_char = ProductChar(
                    char_name_id = char,
                    parent = product,
                    char_value = value
                )
                product_char.save()
                char_count += 1

            old_char_id = request.POST.getlist('old_char_id')
            old_char_name = request.POST.getlist('old_text_name')
            old_char_value = request.POST.getlist('old_char_value')
            old_char_count = 0

            for id in old_char_id:

                old_char = ProductChar.objects.get(id=id)
                old_char.char_name_id = old_char_name[old_char_count]
                old_char.char_value = old_char_value[old_char_count]
                
                old_char.save()
                old_char_count += 1
            
            # Изменение старых опций
            

            
        
            # Изменение старых опций
            old_id = request.POST.getlist('old_id')
            old_type = request.POST.getlist('old_type')
            old_option_value = request.POST.getlist('old_option_value')
            old_option_weight = request.POST.getlist('old_option_weight')
            old_option_sort = request.POST.getlist('old_option_sort')
            
            old_option_price = request.POST.getlist('old_option_price')
            
            old_image_status = request.POST.getlist('old_image_status')
            old_count = 0
            for old in old_id:
                sort = old_option_sort[old_count]
                # print('SORT', sort)
                if sort == '':
                    sort = 0
                price = old_option_price[old_count]
                price = price.replace(',', '.')
                old_option = ProductOption.objects.get(id=old)
                old_option.type_id = old_type[old_count]
                old_option.option_weight = old_option_weight[old_count]
                old_option.sort = sort
                old_option.option_value = old_option_value[old_count]
               
                old_option.option_price = price
               
                old_option.image_status = old_image_status[old_count]
                old_option.save()
                try:
                    old_images = request.FILES.getlist('option_images_old-'+str(old_count))
                    for im in old_images:
                        image = OptionImage(src=im, parent=old_option)
                        image.save()
                except Exception as e:
                    print(e)



                old_count += 1

            # Опции
            options = request.POST.getlist('type')
            option_weight = request.POST.getlist('option_weight')
            option_value = request.POST.getlist('option_value')
            option_sort = request.POST.getlist('option_sort')
            option_price = request.POST.getlist('option_price')
            
            image_status = request.POST.getlist('image_status')
            o_count = 0
            for option in options:
                sort = option_sort[o_count]
                if sort == '':
                    sort = 0
                opt = ProductOption(
                    parent = product,
                    type_id = option,
                    option_weight = option_weight[o_count],
                    option_value = option_value[o_count],
                    sort = option_sort[o_count],
                    option_price = Decimal(option_price[o_count]),
                   
                    image_status = image_status[o_count],
                )
                opt.save()
                
                try:
                    images_name = 'option_images-'+str(o_count)
                    option_images = request.FILES.getlist(images_name)
                    for image in option_images:
                        o_image = OptionImage(parent=opt, src=image)
                        o_image.save()
                except:
                    pass

                o_count += 1

            return redirect('admin_product')
        else:
            return render(request, 'shop/product/product_edit.html', {'form': form})

    context = {
        'form': form,
        'option_form': option_form,
        'product_char_form': product_char_form,
        'options': options,
        'all_options': all_options,
        'all_chars': all_chars,
        'chars': chars,
        'images': images,
        'product': product
    }

    return render(request, 'shop/product/product_edit.html', context)
    

from django.db import transaction
from django.apps import apps

@check_user_rights(['add_products'])
def product_save_as(request, pk):
    
    
    original_product = Product.objects.get(id=pk)
    options = ProductOption.objects.filter(parent_id=pk)
    chars = ProductChar.objects.filter(parent_id=pk)
    images = ProductImage.objects.filter(parent_id=pk)


    # Создаем новый объект Product с атрибутами из существующего объекта
    new_product = Product()

    # Копируем значения всех полей из существующего объекта в новый объект
    for field in original_product._meta.fields:
        setattr(new_product, field.name, getattr(original_product, field.name))

    # Очищаем ID, чтобы создать новую запись в базе данных
    new_product.id = None
    new_product.thumb = None
    new_product.external_id = None
    # Обновляем slug или другие уникальные поля, если это необходимо
    new_product.slug = get_unique_slug(original_product.slug)
    # Сохраняем новый объект Product
    new_product.save()

    
    for option in options:
        new_option = ProductOption()
        for field in option._meta.fields:
            setattr(new_option, field.name, getattr(option, field.name))
        new_option.parent = new_product
        new_option.pk = None  # Указываем Django создать новый объект
        new_option.save()

    for char in chars:
        new_char = ProductChar()
        for field in char._meta.fields:
            setattr(new_char, field.name, getattr(char, field.name))
        new_char.parent = new_product
        new_char.pk = None  # Указываем Django создать новый объект
        new_char.save()

    for image in images:
        new_image = ProductImage()
        for field in image._meta.fields:
            setattr(new_image, field.name, getattr(image, field.name))
        new_image.parent = new_product
        new_image.pk = None  # Указываем Django создать новый объект

        new_image.save()



    return redirect('product_edit', pk=new_product.pk)



    

def get_unique_slug(slug):
    # Функция, чтобы обеспечить уникальность slug
    # Возможно, вам потребуется настроить его для вашей модели
    # Например, добавить случайные символы к slug, если он уже существует
    unique_slug = slug
    counter = 1
    while Product.objects.filter(slug=unique_slug).exists():
        unique_slug = f"{slug}-{counter}"
        counter += 1
    return unique_slug



@check_user_rights(['add_products'])
def product_delete(request, pk):

    product = Product.objects.get(id=pk)
    product.delete()


    return redirect('admin_product')

@check_user_rights(['add_products'])
def stop_list(request, pk):
    product = Product.objects.get(id=pk)

    if product.status == True:
        product.status = False
        product.save()

    else:
        product.status = True
        product.save()

    return redirect('admin_product')

@check_user_rights(['add_products'])
def change_new(request, pk):
    product = Product.objects.get(id=pk)
    if product.new == True:
        product.new = False
        product.save()
    else:
        product.new = True
        product.save()

    return redirect('admin_product')

@check_user_rights(['add_products'])
def change_hit(request, pk):
    product = Product.objects.get(id=pk)
    if product.bestseller == True:
        product.bestseller = False
        product.save()
    else:
        product.bestseller = True
        product.save()

    return redirect('admin_product')

@check_user_rights(['add_products'])
def change_price(request, pk):
    product = Product.objects.get(id=pk)
    if request.method == 'POST':
        price = request.POST.get('price')
        product.price = price
        product.save()
        return redirect('admin_product')
    

@check_user_rights(['add_products'])
def change_old_price(request, pk):
    product = Product.objects.get(id=pk)
    
    if request.method == 'POST':
        old_price = request.POST.get('price')
        # print(old_price)
        if old_price == '':
            old_price = None
        product.old_price = old_price
        product.save()
        return redirect('admin_product')
    

@check_user_rights(['add_products'])
def product_in_cart(request, pk):
    product = Product.objects.get(id=pk)
    if product.in_cart == True:
        product.in_cart = False
        product.save()
    else:
        product.in_cart = True
        product.save()

    return redirect('admin_product')


@check_user_rights(['add_products'])
def product_image_delete(request, pk):

    image = ProductImage.objects.get(id=pk)
    image.delete()

    return redirect('admin_product')


@check_user_rights(['add_products'])
def product_char_delete(request, pk):

    product_char = ProductChar.objects.get(id=pk)
    product_char.delete()

    return redirect('admin_product')

@check_user_rights(['add_products'])
def option_delete(request, pk):
        
    option = ProductOption.objects.get(id=pk)
    option.delete()

    return redirect('admin_product')

@check_user_rights(['add_products'])
def option_image_delete(request, pk):

    image = OptionImage.objects.get(id=pk)
    image.delete()

    parent_id = image.parent.id

    option = ProductOption.objects.get(id=parent_id)
    images = option.images.all()
    
    if images.count() == 0:
        option.image_status = False
        option.save()

    return redirect('admin_product')




# !!! Сопутствующие товары !!!




@check_user_rights(['add_relateds'])
def related(request):

    context = {
        'relateds': Product.objects.filter(related=True)
    }

    return render(request, 'shop/related/related.html', context)


@check_user_rights(['add_relateds'])
def related_add(request):

    if request.method == 'POST':
        form = RelatedProductsForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('related')

        else:
            return render(request, 'shop/related/related_add.html', {'form': form})

    form = RelatedProductsForm(
        {
        'related': True,
        'all_cats': True,
        'free': 0,
        'minimum': 1,
        'price': 0,
        
        }
    )
    context = {
        'form': form,
    }

    return render(request, 'shop/related/related_add.html', context)

@check_user_rights(['add_relateds'])
def related_edit(request, pk):
    related = Product.objects.get(id=pk)
    if request.method == 'POST':
        form = RelatedProductsForm(request.POST, request.FILES, instance=related)
        if form.is_valid():
            form.save()
            return redirect('related')
        else:
            return render(request, 'shop/related/related_edit.html', {'form': form})
            
    form = RelatedProductsForm(instance=related)
    context = {
        'form': form,
    }
    return render(request, 'shop/related/related_edit.html', context)


@check_user_rights(['add_relateds'])
def related_delete(request, pk):
    related = Product.objects.get(id=pk)
    related.delete()
    return redirect('related')
    



# !!! Сопутствующие товары !!!

from shop.models import Combo, ComboItem

# !!! Комбо !!!
@check_user_rights(['add_combos'])
def admin_combo(request):

    combos = Combo.objects.all()    
    context = {
        'combos': combos
    }

    return render(request, 'shop/combo/admin_combo.html', context)




@check_user_rights(['add_combos'])
def add_combo(request):

    if request.method == 'POST':
        
        form = ComboForm(request.POST, request.FILES)

        if form.is_valid():
            combo = form.save(commit=False)
        
            items = request.POST.getlist('item')
            item_prices = request.POST.getlist('item_price')

            
            combo.save()

            cat_list = []
            for item in items:
                product = Product.objects.get(name=item)
                if product.parent.name not in cat_list:
                    cat_list.append(product.parent.name)
            

            for cat in cat_list:

                count = 0
                for item in items:
                    product = Product.objects.get(name=item)

                    if product.parent.name == cat:
                        ComboItem.objects.create(
                            combo=combo,
                            product=product,
                            cat=product.parent.name,
                            price=item_prices[count]
                        )
                    count += 1

            return redirect('admin_combo')
        
        else:
            return render(request, 'shop/combo/edit_combo.html', {'form':form})
    
    products = Product.objects.all().exclude(related=True)


    context = {
        'products': products
    }

    return render(request, 'shop/combo/add_combo.html', context)



@check_user_rights(['add_combos'])
def edit_combo(request, pk):
    combo = Combo.objects.get(id=pk)

    if request.method == 'POST':

        form = ComboForm(request.POST, request.FILES, instance=combo)

        if form.is_valid():
            combo = form.save(commit=False)
        
            items = request.POST.getlist('item')
            item_prices = request.POST.getlist('item_price')

            
            combo.save()

            items_combo = ComboItem.objects.filter(combo=combo)
            items_combo.delete()

            cat_list = []
            for item in items:
                product = Product.objects.get(name=item)
                if product.parent.name not in cat_list:
                    cat_list.append(product.parent.name)
            

            for cat in cat_list:

                count = 0
                for item in items:
                    product = Product.objects.get(name=item)

                    if product.parent.name == cat:
                        item_price = item_prices[count].replace(',','.')
                        ComboItem.objects.create(
                            combo=combo,
                            product=product,
                            cat=product.parent.name,
                            price=item_price
                        )
                    count += 1

            return redirect('admin_combo')
        
        else:
            return render(request, 'shop/combo/edit_combo.html', {'form':form})
    

    
    form = ComboForm(instance=combo)
    products = Product.objects.all()
    context = {
        'products': products,
        'combo':combo,
        'form': form
    }

    return render(request, 'shop/combo/edit_combo.html', context)




@check_user_rights(['add_combos'])
def delete_combo(request, pk):
    combo = Combo.objects.get(id=pk)
    combo.delete()

    return redirect('admin_combo')


@check_user_rights(['add_combos'])
def delete_combo_item(request, pk):
    
    item = ComboItem.objects.get(id=pk)
    parent_id = item.combo.id

    item.delete()

    return redirect(f'/admin/edit_combo/{parent_id}/')




# !!! Комбо !!!



# !!! БЛОГ !!!
@check_user_rights(['add_posts'])
def blog_settings(request):
    blog_settings = BlogSetup.objects.get()
    if request.method == 'POST':
        form = BlogSetupForm(request.POST, instance=blog_settings)
        if form.is_valid():
            form.save()
            return redirect('blog_settings')
        else:
            return render(request, 'blog/blog_settings.html', {'form':form})

    form = BlogSetupForm(instance=blog_settings)
    context = {
        'form': form
    }

    return render(request, 'blog/blog_settings.html', context)




@check_user_rights(['add_posts'])
def blog_category(request):
    q = request.GET.get('q')
    sort = request.GET.getlist('sort')
    try:
        sort_t = sort[0]
    except:
        sort_t = sort
    try:
        blog_category = BlogCategory.objects.filter(Q(name__icontains=q)).order_by(*sort)
    except:
        blog_category = BlogCategory.objects.all().order_by(*sort)
    context = {
        # Разрешить поиск на странице
        'search': 'search',
        'blog_category': blog_category,
        'q': q,
        'sort': sort_t,
    }
    return render(request, 'blog/blog_category/blog_category.html', context)


@check_user_rights(['add_posts'])
def blog_category_add(request):
    if request.method == 'POST':
        form = BlogCategoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('blog_category')
        else:
            return render(request, 'blog/blog_category/blog_category_add.html', {'form': form})
    form = BlogCategoryForm()
    context = {
        'form':form,
    } 
    return render(request, 'blog/blog_category/blog_category_add.html', context)

@check_user_rights(['add_posts'])
def blog_category_edit(request, pk):
    blog_category = BlogCategory.objects.get(id=pk)
    if request.method == 'POST':
        form = BlogCategoryForm(request.POST, request.FILES, instance=blog_category)
        if form.is_valid():
            form.save()
            return redirect('blog_category')
        else:
            return render(request, 'blog/blog_category/blog_category_edit.html', {'form': form})
    form = BlogCategoryForm(instance=blog_category)
    context = {
        'form':form,
    } 
    return render(request, 'blog/blog_category/blog_category_edit.html', context)

@check_user_rights(['add_posts'])
def blog_category_delete(request, pk):
    blog_category = BlogCategory.objects.get(id=pk)
    blog_category.delete()
    return redirect('blog_category')


@check_user_rights(['add_posts'])
def blog_post(request):
    q = request.GET.get('q')
    sort = request.GET.getlist('sort')
    try:
        sort_t = sort[0]
    except:
        sort_t = sort
    try:
        posts = Post.objects.filter(Q(name__icontains=q)).order_by(*sort)
    except:
        posts = Post.objects.all().order_by(*sort).order_by('-id')
    context = {
        # Разрешить поиск на странице
        'search': 'search',
        'posts': posts,
        'q': q,
        'sort': sort_t,
    }


    return render(request, 'blog/blog_post/blog_post.html', context)


@check_user_rights(['add_posts'])
def post_add(request):
    try:
        post = Post.objects.get(draft=True)
    except:
        post = Post(draft=True)
        post.save()

    form = PostForm(instance=post)
    block_form = PostBlockForm()
    if request.method == 'POST':
        post.draft = False
        post.published = True
        form = PostForm(request.POST, request.FILES, instance=post)
        if form.is_valid():
            form.save()

            return redirect('blog_post')
    context = {
        'post': post,
        'form': form,
        'block_form': block_form,
    }


    return render(request, 'blog/blog_post/post_add.html', context)

@check_user_rights(['add_posts'])
def post_edit(request, pk):
    
    post = Post.objects.get(id=pk)
    
    form = PostForm(instance=post)
    block_form = PostBlockForm()
    if request.method == 'POST':
        post.draft = False
        post.published = True
        form = PostForm(request.POST, request.FILES, instance=post)
        if form.is_valid():
            form.save()

            return redirect('blog_post')
    context = {
        'post': post,
        'form': form,
        'block_form': block_form,
    }


    return render(request, 'blog/blog_post/post_edit.html', context)

@check_user_rights(['add_posts'])
def post_delete(request, pk):
    post = Post.objects.get(id=pk)
    post.delete()
    return redirect('blog_post')


@check_user_rights(['add_posts'])
def post_draft(request):

    if request.method == 'POST':
        post = Post.objects.get(draft=True)
        form = PostForm(request.POST, request.FILES, instance=post)
        if form.is_valid():
            form.save() 

            return redirect('blog_post')

        else:

            return render(request, 'blog/blog_post/post_add.html', {'form': form})

@check_user_rights(['add_posts'])
def post_block(request):
    if request.method == 'POST':
        parent = request.POST['parent']
        type = request.POST['type']
        order = request.POST['order']
        next = request.POST['next']
        try:
            text = request.POST['text']
            block = PostBlock(parent_id=parent, text=text, type=type, order=order)
            block.save()
        except:
            pass
        try:
            title = request.POST['title']
            block = PostBlock(parent_id=parent, title=title, type=type, order=order)
            block.save()
        except:
            pass
        try:
            image = request.FILES['image']
            block = PostBlock(parent_id=parent, image=image, type=type, order=order)
            block.save()
        except:
            pass
        try:
            video = request.FILES['video']
            block = PostBlock(parent_id=parent, video=video, type=type, order=order)
            block.save()
        except:
            pass

        
        return redirect(next)


@check_user_rights(['add_posts'])
def post_block_edit(request, pk):
    block = PostBlock.objects.get(id=pk)
    next = request.POST['next']
    if request.method == 'POST':
        try:
            text = request.POST['text']
            block.text = text
            block.save()
        except:
            pass
        try:
            title = request.POST['title']
            block.title = title
            block.save()
        except:
            pass
        try:
            image = request.FILES['image']
            block.image = image
            block.save()
        except:
            pass
        try:
            video = request.FILES['video']
            block.video = video
            block.save()
        except:
            pass

        
        return redirect(next)


@check_user_rights(['add_posts'])
def post_block_edit_delete(request, pk):

    block = PostBlock.objects.get(id=pk)
    block.delete()

    return redirect('post_edit', block.parent.id)


@check_user_rights(['add_posts'])
def post_block_add_delete(request, pk):

    block = PostBlock.objects.get(id=pk)
    block.delete()

    return redirect('post_add')





# !!! СТАТИКА !!!


@check_user_rights(['add_sliders'])
def admin_slider(request):
    sliders = Slider.objects.all().order_by('order')
    try:
        slider_setup = SliderSetup.objects.get()
    except:
        slider_setup = SliderSetup()
        slider_setup.save()

    if request.method == 'POST':
        form = SliderSetupForm(request.POST, instance=slider_setup)
        if form.is_valid():
            form.save()
            return redirect('admin_slider')
        else:
            return render(request, 'static/slider.html', {'form': form})

    setup_form = SliderSetupForm(instance=slider_setup)
    context = {
        'setup_form': setup_form,
        'sliders': sliders
    }
    return render(request, 'static/slider.html', context)

@check_user_rights(['add_sliders'])
def slider_add(request):

    if request.method == 'POST':
        form = SliderForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('admin_slider')

        else:
            return render(request, 'static/slider_add.html', {'form': form})
    form = SliderForm()
    context = {
        'form': form,
    }

    return render(request, 'static/slider_add.html', context)



@check_user_rights(['add_sliders'])
def slider_edit(request, pk):
    slider = Slider.objects.get(id=pk)
    if request.method == 'POST':
        form = SliderForm(request.POST, request.FILES, instance=slider)
        if form.is_valid():
            form.save()
            return redirect('admin_slider')

        else:
            return render(request, 'static/slider_edit.html', {'form': form})
    form = SliderForm(instance=slider)
    context = {
        'form': form,
    }

    return render(request, 'static/slider_edit.html', context)


@check_user_rights(['add_sliders'])
def slider_delete(request, pk):
    slider = Slider.objects.get(id=pk)
    slider.delete()
    return redirect('admin_slider')


@check_user_rights(['add_statics'])
def admin_pages(request):

    context = {
        'pages': Page.objects.all()
    }

    return render(request, 'static/admin_pages.html', context)


@check_user_rights(['add_statics'])
def page_add(request):

    if request.method == 'POST':
        form = PageForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('admin_pages')

        else:
            return render(request, 'static/page_add.html', {'form': form})

    form = PageForm()
    context = {
        'form': form,
    }

    return render(request, 'static/page_add.html', context)

@check_user_rights(['add_statics'])
def page_edit(request, pk):
    page = Page.objects.get(id=pk)
    if request.method == 'POST':
        form = PageForm(request.POST, request.FILES, instance=page)
        if form.is_valid():
            form.save()
            return redirect('admin_pages')
        else:
            return render(request, 'static/page_edit.html', {'form': form})
            
    form = PageForm(instance=page)
    context = {
        'form': form,
    }
    return render(request, 'static/page_edit.html', context)


@check_user_rights(['add_statics'])
def page_delete(request, pk):
    page = Page.objects.get(id=pk)
    page.delete()
    return redirect('admin_pages')

@check_user_rights(['add_statics'])
def page_item_add(request, pk):
    page = Page.objects.get(id=pk)
    if request.method == 'POST':
        form = PageItemForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('admin_pages')
        else:
            return render(request, 'static/page_item_add.html', {'form': form})

    form = PageItemForm({
        'page': page
    })
    context = {
        'form': form,
        'page': page
    }

    return render(request, 'static/page_item_add.html', context)


@check_user_rights(['add_statics'])
def page_item_edit(request, pk):
    page_item = PageItem.objects.get(id=pk)
    if request.method == 'POST':
        form = PageItemForm(request.POST, request.FILES, instance=page_item)
        if form.is_valid():
            form.save()
            return redirect('admin_pages')
        else:
            return render(request, 'static/page_item_add.html', {'form': form})
            
    form = PageItemForm(instance=page_item)
    context = {
        'form': form,
    }
    return render(request, 'static/page_item_add.html', context)


@check_user_rights(['add_statics'])
def page_item_delete(request, pk):
    page_item = PageItem.objects.get(id=pk)
    page_item.delete()
    return redirect('admin_pages')

@check_user_rights(['add_statics'])
def admin_images(request):


    context = {
        'images': PlaceImages.objects.all()
    }

    return render(request, 'static/admin_images.html', context)


@check_user_rights(['add_statics'])
def image_add(request):

    if request.method == 'POST':
        form = ImageForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('admin_images')
    
        else:
            return render(request, 'static/image_add.html', {'form': form})
        
    form = ImageForm()

    context = {
        'form': form
    }

    return render(request, 'static/image_add.html', context)

        
@check_user_rights(['add_statics'])
def image_edit(request, pk):
    image = PlaceImages.objects.get(id=pk)
    if request.method == 'POST':
        form = ImageForm(request.POST, request.FILES, instance=image)
        if form.is_valid():
            form.save()
            return redirect('admin_images')
        else:
            return render(request, 'static/image_add.html', {'form': form})
            
    form = ImageForm(instance=image)
    context = {
        'form': form,
    }
    return render(request, 'static/image_add.html', context)

@check_user_rights(['add_statics'])
def image_delete(request, pk):
    image = PlaceImages.objects.get(id=pk)
    image.delete()
    return redirect('admin_images')

# !!! СТАТИКА !!!


# !!! Пользователи USERS !!!

from itertools import chain
@check_user_rights(['view_customers'])
def admin_users(request):
    q = request.GET.get('q')
    sort = request.GET.getlist('sort')

    try:
        sort_t = sort[0]
    except:
        sort_t = sort

    users = User.objects.all().exclude(is_staff=True)
    try:
        users_pr = UserProfile.objects.filter(Q(phone__icontains=q)).order_by(*sort)
    except:
        try:
            users_pr = UserProfile.objects.all().order_by(*sort)
        except:
            users_pr = UserProfile.objects.all()


    # Исключаем персонал и складываем две разные фильтрации
    users_pr = list(chain(users_pr.filter(user=None), UserProfile.objects.filter(user__in=users)))

    context = {
        'users': users,
        'users_pr': users_pr,
        'search': 'search',
        'q': q,
        'sort': sort_t,
    }

    return render(request, 'users/admin_users.html', context)


@check_user_rights(['view_customers'])
def users_detail(request, pk):

    user_pr = UserProfile.objects.get(id=pk)

    context = {
      
        'user_pr': user_pr,
       
    }

   
    return render(request, 'users/users_detail.html', context)

@check_user_rights(['view_customers'])
def users_delete(request, pk):

    user = UserProfile.objects.get(id=pk)
    user.delete()

    return redirect('admin_users')


# !!! Пользователи USERS !!!







import zipfile
from django.core.files.storage import FileSystemStorage
import os
import shutil
from openpyxl import load_workbook
from pytils.translit import slugify
from PIL import Image


from django.core.files.images import ImageFile

@check_user_rights(['upload_csv'])
def product_upload(request):

    
    prod = Product.objects.all()

    
    prod.delete()
    
    # file = 'csv_upload/upload/upload.xlsx'
    file = 'csv_upload/upload.xlsx'
    
    wb = load_workbook(file)

    
    ws = wb.active
    count = 0
    connect = 0
    for row in ws.iter_rows(min_row=1):
        if count == 0:
            pass
        else:
            name = row[1].value
            slug = slugify(name)
           
            cat_name = row[2].value
            cat_slug = slugify(cat_name)
            
            price = row[3].value
            old_price = row[4].value

            if old_price == '':
                old_price = None

            weight = row[5].value

            if weight == '':
                weight = None
            
            short_description = row[6].value
            if short_description == '':
                short_description = None
           
            
            
            # filename = str(str(row[7].value).split('.')[0])

            # thumb = 'csv_upload/' + filename

            # files = os.listdir('csv_upload/')

            # if any(filename in f for f in files):
            #     print(f"Изображение {filename} найдено")
            #     formats = ['.jpg', ' .jpg', '.jpeg', ' .jpeg', '.webp', ' .webp', '.png', ' .png']
            #     format = None
            #     for ext in formats:
            #         try:
            #             with open('csv_upload/'+filename + ext, "rb") as f:
            #                 format = ext.strip()
            #                 break
            #         except:
            #             pass

            #     if format:
            #         print("Формат файла: " + format)
            #     else:
            #         print("Не удалось определить формат файла")

            # else:
            #     print(f"Изображение {name} не найдено")

            
            
            try:
                # thumb_path = 'csv_upload/'
                # thumb_for_format = thumb_path + filename + format
                # thumb_file = open(thumb_for_format, 'rb')
                # thumb_image = ImageFile(thumb_file)

            
                try:
                    cat = Category.objects.get(slug=cat_slug)

                except:
                    cat = Category.objects.create(
                        name=cat_name,
                        slug=cat_slug,
                        top=True
                        )

                try:
                    product = Product.objects.get(slug=slug)
  
                    product.stock=1
                    product.price=price
                    product.old_price=old_price
                    
                    # product.thumb=thumb_image
                    product.parent=cat
                    product.weight=weight
                    product.short_description=short_description

                    product.save()

                except:
                    product = Product.objects.create(
                        name=name,
                        slug=slug,
                        
                        stock=1,
                        price=price,
                        old_price=old_price,
                      
                        # thumb=thumb_image,
                        parent=cat,
                        weight=weight,
                        short_description=short_description
                        )
            except Exception as e:
                print(e)

        count += 1
                
# product_upload()


                
# !!!! Загрузка csv

@check_user_rights(['upload_csv'])
def csv_upload(request):

    if request.method == 'POST':
        file = request.FILES['file']

        folder='csv_upload/' 

        if not os.path.exists(folder):
            os.makedirs(folder)
        else:
            shutil.rmtree('csv_upload')
        
        fs = FileSystemStorage(location=folder) #defaults to   MEDIA_ROOT  
        filename = fs.save(file.name, file)
        file_url = fs.url(filename)
        
        file_to_open = f'csv_upload/{file}'
        fantasy_zip = zipfile.ZipFile(file_to_open)
        fantasy_zip.extractall('csv_upload')
        fantasy_zip.close()


        os.remove(file_to_open)

        product_upload(request)

        shutil.rmtree('csv_upload')

        return redirect('csv_upload')



    return render(request, 'upload/csv_upload.html')




# !!!! Загрузка csv







@check_user_rights(['change_settings'])
def admin_subdomain(request):

    context = {
        'subdomains': Subdomain.objects.all()
    }

    return render(request, 'subdomain/admin_subdomain.html', context)


@check_user_rights(['change_settings'])
def add_subdomain(request):

    if request.method == 'POST':
        form = SubdomainsForm(request.POST, request.FILES)

        if form.is_valid():
            form.save()

            return redirect('admin_subdomain')

        else:
            return render(request, 'subdomain/add_subdomain.html', {'form': form})




    form = SubdomainsForm()

    context = {
        'form' : form
    }

    return render(request, 'subdomain/add_subdomain.html', context)


@check_user_rights(['change_settings'])
def edit_subdomain(request, pk):

    subdomain = Subdomain.objects.get(id=pk)

    if request.method == 'POST':
        form = SubdomainsForm(request.POST, request.FILES, instance=subdomain)

        if form.is_valid():
            form.save()

            return redirect('admin_subdomain')

        else:
            return render(request, 'subdomain/edit_subdomain.html', {'form': form})




    form = SubdomainsForm(instance=subdomain)

    context = {
        'form' : form
    }

    return render(request, 'subdomain/edit_subdomain.html', context)


@check_user_rights(['change_settings'])
def delete_subdomain(request, pk):

    subdomain = Subdomain.objects.get(id=pk)
    subdomain.delete()

    return redirect('admin_subdomain')



# !!! Субдомены !!!



# !!! Интеграции !!!

@check_user_rights(['add_integrations'])
def integration(request):
    try:
        integrations = Integrations.objects.get()
    except:
        integrations = None
    
    context = {
        'integrations': integrations
    }

    return render(request, 'integration/integration.html', context)



@check_user_rights(['add_integrations'])
def add_integration(request):

    if request.method == 'POST':
        form = IntegrationsForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('integration')
        else:
            return render(request, 'integration/add_integration.html', {'form': form})

    form = IntegrationsForm()

    context = {
        'form': form
    }
    return render(request, 'integration/add_integration.html', context)


@check_user_rights(['add_integrations'])
def edit_integration(request, pk):
    integration = Integrations.objects.get(id=pk)
    if request.method == 'POST':
        form = IntegrationsForm(request.POST, request.FILES, instance=integration)
        if form.is_valid():
            form.save()
            return redirect('integration')
        else:
            return render(request, 'integration/add_integration.html', {'form': form})

    form = IntegrationsForm(instance=integration)
    context = {
        'form': form
    }
    return render(request, 'integration/add_integration.html', context)


@check_user_rights(['add_integrations'])
def delete_integration(request, pk):
    integration = Integrations.objects.get(id=pk)
    integration.delete()
    return redirect('integration')


from integrations.iiko import load_menu
@check_user_rights(['add_integrations'])
def catalogs_synch(request):

    if request.method == 'POST':

        clean = request.POST.get('clean')
        product_clean = request.POST.get('product_clean')

        if clean == 'on':
            status_clean = True
        else:
            status_clean = False

        if product_clean == 'on':
            product_status = True
        else:
            product_status = False


      
        load_menu(status_clean, product_status)
      
        

        return redirect('integration')

@check_user_rights(['add_integrations'])
def synch_cron(request):


    return redirect('integration')

# !!! Интеграции !!!

from integrations.cron import *




# !!! Настройка рабочих дней и времени доставки !!!

@check_user_rights(['change_shop_settings'])
def add_worksday(request):
    if request.method == 'POST':
        form = WorksdayForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('shop_settings')
        else:
            return render(request, 'shop/works_day/add_worksday.html', {'form': form})

    form = WorksdayForm()

    context = {
        'form': form
    }

    return render(request, 'shop/works_day/add_worksday.html', context)


@check_user_rights(['change_shop_settings'])
def edit_worksday(request, pk):
    worksday = WorkDay.objects.get(id=pk)
    if request.method == 'POST':
        form = WorksdayForm(request.POST, instance=worksday)
        if form.is_valid():
            form.save()
            return redirect('shop_settings')
        else:
            return render(request, 'shop/works_day/edit_worksday.html', {'form': form})

    form = WorksdayForm(instance=worksday)

    context = {
        'form': form
    }

    return render(request, 'shop/works_day/edit_worksday.html', context)


@check_user_rights(['change_shop_settings'])
def delete_worksday(request, pk):
    worksday = WorkDay.objects.get(id=pk)
    worksday.delete()
    return redirect('shop_settings')
# !!! Настройка рабочих дней и времени доставки !!!



# !!! ДОСТАВКА !!!
@check_user_rights(['add_delivery_service'])
def admin_delivery(request):

    try:
        deliverys = Delivery.objects.get()
    except:
        deliverys = None


    context = {
        'deliverys': deliverys
    }
    return render(request, 'delivery/admin_delivery.html', context)


@check_user_rights(['add_delivery_service'])
def delivery_add(request):

    if request.method == 'POST':
        form = DeliveryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('admin_delivery')
        else:
            return render(request, 'delivery/add_delivery.html', {'form': form})

    form = DeliveryForm()

    context = {
        'form': form
    }

    return render(request, 'delivery/add_delivery.html', context)


@check_user_rights(['add_delivery_service'])
def delivery_edit(request, pk):
    delivery = Delivery.objects.get(id=pk)
    if request.method == 'POST':
        form = DeliveryForm(request.POST, instance=delivery)
        if form.is_valid():
            form.save()
            subprocess.call(["touch", RESET_FILE])
            return redirect('admin_delivery')
        else:
            return render(request, 'delivery/add_delivery.html', {'form': form})

    form = DeliveryForm(instance=delivery)

    context = {
        'form': form
    }

    return render(request, 'delivery/add_delivery.html', context)


@check_user_rights(['add_delivery_service'])
def delivery_delete(request, pk):
    delivery = Delivery.objects.get(id=pk)
    delivery.delete()
    return redirect('admin_delivery')




@check_user_rights(['add_constructors'])
def admin_food_constructor(request):

    food_constructors = FoodConstructor.objects.all()

    context = {
        'food_constructors': food_constructors
    }
    return render(request, 'shop/food_constructor/food_constructor.html', context)



@check_user_rights(['add_constructors'])
def add_food_constructor(request):

    if request.method == 'POST':
        form = FoodConstructorForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('admin_food_constructor')
        else:
            return render(request, 'shop/food_constructor/constructor/add_food_constructor.html', {'form': form})

    form = FoodConstructorForm()

    context = {
        'form': form
    }

    return render(request, 'shop/food_constructor/constructor/add_food_constructor.html', context)


@check_user_rights(['add_constructors'])
def edit_food_constructor(request, pk):
    food_constructor = FoodConstructor.objects.get(id=pk)
    if request.method == 'POST':
        form = FoodConstructorForm(request.POST, request.FILES, instance=food_constructor)
        if form.is_valid():
            form.save()
            return redirect('admin_food_constructor')
        else:
            return render(request, 'shop/food_constructor/constructor/add_food_constructor.html', {'form': form})

    form = FoodConstructorForm(instance=food_constructor)

    context = {
        'form': form
    }

    return render(request, 'shop/food_constructor/constructor/add_food_constructor.html', context)

@check_user_rights(['add_constructors'])
def delete_food_constructor(request, pk):
    food_constructor = FoodConstructor.objects.get(id=pk)
    food_constructor.delete()
    return redirect('admin_food_constructor')


@check_user_rights(['add_constructors'])
def add_constructor_category(request, pk):

    parent = FoodConstructor.objects.get(id=pk)

    if request.method == 'POST':
        form = ConstructorCategoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('admin_food_constructor')
        else:
            return render(request, 'shop/food_constructor/category/add_constructor_category.html', {'form': form})

    form = ConstructorCategoryForm({
        'parent': parent
    })

    context = {
        'form': form
    }

    return render(request, 'shop/food_constructor/category/add_constructor_category.html', context)


@check_user_rights(['add_constructors'])
def edit_constructor_category(request, pk):

    category = ConstructorCategory.objects.get(id=pk)

    if request.method == 'POST':
        form = ConstructorCategoryForm(request.POST, request.FILES, instance=category)
        if form.is_valid():
            form.save()
            return redirect('admin_food_constructor')
        else:
            return render(request, 'shop/food_constructor/category/add_constructor_category.html', {'form': form})

    form = ConstructorCategoryForm(instance=category)

    context = {
        'form': form
    }

    return render(request, 'shop/food_constructor/category/add_constructor_category.html', context)

@check_user_rights(['add_constructors'])
def delete_constructor_category(request, pk):
    category = ConstructorCategory.objects.get(id=pk)
    category.delete()
    return redirect('admin_food_constructor')



@check_user_rights(['add_constructors'])
def add_ingridients(request, pk):

    parent = ConstructorCategory.objects.get(id=pk)

    if request.method == 'POST':
        form = IngridientsForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('admin_food_constructor')
        else:
            return render(request, 'shop/food_constructor/ingridients/add_ingridients.html', {'form': form})

    form = IngridientsForm({
        'parent': parent
    })

    context = {
        'form': form
    }

    return render(request, 'shop/food_constructor/ingridients/add_ingridients.html', context)


@check_user_rights(['add_constructors'])
def edit_ingridients(request, pk):

    ingridient = Ingridients.objects.get(id=pk)

    if request.method == 'POST':
        form = IngridientsForm(request.POST, request.FILES, instance=ingridient)
        if form.is_valid():
            form.save()
            return redirect('admin_food_constructor')
        else:
            return render(request, 'shop/food_constructor/ingridients/add_ingridients.html', {'form': form})

    form = IngridientsForm(instance=ingridient)

    context = {
        'form': form
    }

    return render(request, 'shop/food_constructor/ingridients/add_ingridients.html', context)

@check_user_rights(['add_constructors'])
def delete_ingridients(request, pk):
    ingridient = Ingridients.objects.get(id=pk)
    ingridient.delete()
    return redirect('admin_food_constructor')




@check_user_rights(['change_time_sales'])
def delivery_time_price(request):

    context = {
        'delivery_times': DeliveryTimePrice.objects.all()
    }

    return render(request, 'shop/delivery_time_price/delivery_time_price.html', context)


@check_user_rights(['change_time_sales'])
def add_delivery_time_price(request):

    if request.method == 'POST':
        form = DeliveryTimePriceForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('delivery_time_price')
        else:
            return render(request, 'shop/delivery_time_price/add_delivery_time_price.html', {'form': form})

    form = DeliveryTimePriceForm()

    context = {
        'form': form
    }

    return render(request, 'shop/delivery_time_price/add_delivery_time_price.html', context)


@check_user_rights(['change_time_sales'])
def edit_delivery_time_price(request, pk):

    delivery_time_price = DeliveryTimePrice.objects.get(id=pk)

    if request.method == 'POST':
        form = DeliveryTimePriceForm(request.POST, request.FILES, instance=delivery_time_price)
        if form.is_valid():
            form.save()
            return redirect('delivery_time_price')
        else:
            return render(request, 'shop/delivery_time_price/add_delivery_time_price.html', {'form': form})

    form = DeliveryTimePriceForm(instance=delivery_time_price)

    context = {
        'form': form
    }

    return render(request, 'shop/delivery_time_price/add_delivery_time_price.html', context)


@check_user_rights(['change_time_sales'])
def delete_delivery_time_price(request, pk):
    delivery_time_price = DeliveryTimePrice.objects.get(id=pk)
    delivery_time_price.delete()
    return redirect('delivery_time_price')





@check_user_rights(['add_reviews'])
def admin_reviews(request):

    context = {
        'reviews': Reviews.objects.all()
    }

    return render(request, 'marketing/reviews/reviews.html', context)


@check_user_rights(['add_reviews'])
def add_reviews(request):

    if request.method == 'POST':
        form = ReviewsForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('admin_reviews')
        else:
            return render(request, 'marketing/reviews/add_reviews.html', {'form': form})

    form = ReviewsForm()

    context = {
        'form': form
    }

    return render(request, 'marketing/reviews/add_reviews.html', context)

@check_user_rights(['add_reviews'])
def edit_reviews(request, pk):

    review = Reviews.objects.get(id=pk)

    if request.method == 'POST':
        form = ReviewsForm(request.POST, request.FILES, instance=review)
        if form.is_valid():
            form.save()
            return redirect('admin_reviews')
        else:
            return render(request, 'marketing/reviews/add_reviews.html', {'form': form})

    form = ReviewsForm(instance=review)

    context = {
        'form': form
    }

    return render(request, 'marketing/reviews/add_reviews.html', context)

@check_user_rights(['add_reviews'])
def delete_reviews(request, pk):
    review = Reviews.objects.get(id=pk)
    review.delete()
    return redirect('admin_reviews')